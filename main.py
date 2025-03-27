# Importación de librerías necesarias para la ejecución del script
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import requests
import os
import time
import urllib.parse
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import re
from config import USUARIX, CONTRASEÑA, DIRECTORIO_DESCARGAS, CURSOS_A_DESCARGAR, CURSOS_A_EXCLUIR  # Importar credenciales y configuración desde config.py
import errno

# Configura el directorio de descargas para el navegador Chrome
def set_download_directory(driver, download_dir):
    # Configuración para habilitar descargas automáticas en Chrome
    driver.command_executor._commands["send_command"] = ("POST", '/session/$sessionId/chromium/send_command')
    params = {'cmd': 'Page.setDownloadBehavior', 'params': {'behavior': 'allow', 'downloadPath': download_dir}}
    driver.execute("send_command", params)

# Agrega cookies del navegador Selenium a una sesión de requests
def add_cookies_to_session(driver, session):
    # Itera sobre las cookies del navegador y las agrega a la sesión de requests
    for cookie in driver.get_cookies():
        session.cookies.set(cookie['name'], cookie['value'])

# Inicia sesión en la plataforma Paideia
def login(driver, usuarix, contraseña):
    # Navega a la página de inicio de sesión
    driver.get('https://pandora.pucp.edu.pe/pucp/login?service=https%3A%2F%2Fpandora.pucp.edu.pe%2Fpucp%2Fidp%2Fprofile%2FSAML2%2FCallback%3Fsrid%3D_737301cfd029b0244abd21131386b320%26entityId%3Dhttps%253A%252F%252Fcursoshistoricospaideia.pucp.edu.pe%252Fcursos%252Fshibboleth')
    # Completa los campos de usuario y contraseña
    driver.find_element(By.NAME, 'username').send_keys(usuarix)
    driver.find_element(By.NAME, 'password').send_keys(contraseña)
    # Haz clic en el botón de acceso
    driver.find_element(By.XPATH, '//input[@value="Acceder"]').click()
    time.sleep(3)  # Espera para que la sesión se cargue completamente

# Obtiene la lista de cursos disponibles en la plataforma
def obtener_cursos(driver):
    # Navega a la página de cursos
    driver.get('https://cursoshistoricospaideia.pucp.edu.pe/cursos/local/paideia_customs/miscursos/')
    # Espera a que los elementos de los cursos estén presentes
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'coursebox')))
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    resultados = []
    # Extrae información de cada curso
    for coursebox in soup.find_all('div', class_='coursebox'):
        course_id = coursebox.get('data-courseid')
        course_name_tag = coursebox.find('h3', class_='coursename').find('a')
        if course_id and course_name_tag:
            resultados.append({
                'courseid': course_id,
                'coursename': course_name_tag.text.strip(),
                'url': course_name_tag['href']
            })
    return resultados

# Crea directorios de manera segura
def create_directory(path):
    try:
        os.makedirs(path, exist_ok=True)
        print(f"Creando directorio: {path}")
    except OSError as e:
        if e.errno == errno.EACCES:
            print(f"Error: No tienes permisos para crear el directorio {path}.")
            raise
        else:
            raise

# Verifica la conexión a internet antes de continuar
def check_internet_connection(url="https://pandora.pucp.edu.pe"):
    try:
        response = requests.get(url, timeout=5)
        if response.status_code == 200:
            print("Conexión a internet verificada.")
        else:
            print(f"Advertencia: Respuesta inesperada del servidor ({response.status_code}).")
    except requests.ConnectionError:
        print("Error: No se pudo conectar a internet. Verifica tu conexión.")
        raise

# Descarga archivos de los cursos y actualiza las URLs en el registro
def descargar_archivos_y_actualizar_urls(driver, resultados):
    registro = []
    for r in resultados:
        # Configuración del directorio de descargas para cada curso
        courseid = r['courseid']
        coursename = r['coursename'].replace('/', '_')  # Reemplaza caracteres no válidos en nombres de carpetas
        course_dir = os.path.join(DIRECTORIO_DESCARGAS or os.getcwd(), 'Descarga_cursos_historicos', coursename)
        create_directory(course_dir)

        # Configura el directorio de descargas para el curso actual
        set_download_directory(driver, course_dir)

        retry_count = 3  # Número de intentos para acceder a los recursos
        while retry_count > 0:
            try:
                driver.get(f'https://cursoshistoricospaideia.pucp.edu.pe/cursos/course/resources.php?id={courseid}')
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'a')))
                break
            except Exception as e:
                print(f"Error al acceder a los recursos del curso {coursename}: {e}")
                retry_count -= 1
                if retry_count == 0:
                    print(f"No se pudo acceder a los recursos del curso {coursename} después de varios intentos.")
                    continue
                time.sleep(5)  # Espera antes de reintentar

        # Procesa los recursos del curso
        soup_resources = BeautifulSoup(driver.page_source, 'html.parser')
        main_content = soup_resources.find('div', {'role': 'main'})

        for link in main_content.find_all('a', href=True):
            href = link['href']
            name = link.text.strip() or 'archivo_sin_nombre'

            if 'resource/view.php?id=' in href:
                try:
                    driver.get(href)
                    redirected_url = driver.current_url
                    print(f"Redireccionado a: {redirected_url}")

                    if 'pluginfile.php' in redirected_url:
                        file_url = redirected_url
                        filename = urllib.parse.unquote(file_url.split('/')[-1])

                        # Descarga el archivo usando requests
                        session = requests.Session()
                        add_cookies_to_session(driver, session)
                        response = session.get(file_url, stream=True)

                        if response.status_code == 200:
                            filepath = os.path.join(course_dir, filename)
                            with open(filepath, 'wb') as f:
                                for chunk in response.iter_content(chunk_size=8192):
                                    f.write(chunk)
                            print(f"Descargado: {filename}")
                        else:
                            print(f"No se pudo descargar: {name} (Status code: {response.status_code})")
                    else:
                        print(f"Redirección sin descarga: {name}")
                except Exception as e:
                    print(f"Error descargando {name}: {e}")

            # Actualiza la URL final en el registro
            final_url = driver.current_url
            registro.append({
                'Curso': coursename,
                'Nombre enlace': name,
                'URL': final_url
            })
    return registro

# Guarda el registro de descargas en un archivo Excel
def guardar_registro_en_excel(df_registro, excel_file):
    # Crea un archivo Excel con formato
    df_registro.to_excel(excel_file, index=False)
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')  # Color amarillo para resaltar
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    alignment = Alignment(vertical='center', wrap_text=True)
    font = Font(size=15)  # Tamaño de fuente aumentado

    # Define patrones para resaltar enlaces no descargables
    patterns = [
        r"cursos/mod/folder/",
        r"pucp\.zoom\.us/rec/",
        r"facebook",
        r"youtube",
        r"zoom\.us",
        r"drive\.google",
        r"accounts\.google"
    ]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        url_cell = row[2]  # La URL está en la tercera columna
        if any(re.search(pattern, url_cell.value) for pattern in patterns) or 'pluginfile.php' not in url_cell.value:
            for cell in row:
                cell.fill = yellow_fill
        for cell in row:
            cell.border = thin_border
            cell.alignment = alignment
            cell.font = font

    # Ajusta el ancho de las columnas y el alto de las filas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 20

    wb.save(excel_file)
    print("Archivo Excel creado y guardado en:", excel_file)

# Función para filtrar cursos basados en la selección del usuario en config.py
def filtrar_cursos(resultados, cursos_a_descargar):
    if not cursos_a_descargar:
        print("No se especificaron cursos en config.py. Se descargarán todos los cursos.")
        return resultados

    seleccionados = [
        curso for curso in resultados
        if curso['coursename'] in cursos_a_descargar
    ]

    if not seleccionados:
        print("No se encontraron cursos que coincidan con la selección en config.py. Saliendo...")
        exit()

    return seleccionados

# Function to generate regex patterns from keywords
def generar_patrones_regex(keywords):
    # Convert each keyword into a case-insensitive regex pattern
    return [rf"(?i).*{re.escape(keyword)}.*" for keyword in keywords]

# Parse comma-separated strings into lists of keywords
def parse_keywords(keywords_string):
    if not keywords_string.strip():
        return []
    return [keyword.strip() for keyword in keywords_string.split(",")]

# Function to filter courses based on keywords in config.py
def filtrar_cursos_por_keywords(resultados, keywords_incluir_string, keywords_excluir_string):
    keywords_incluir = parse_keywords(keywords_incluir_string)
    keywords_excluir = parse_keywords(keywords_excluir_string)

    if not keywords_incluir and not keywords_excluir:
        print("No se especificaron palabras clave en config.py. Se descargarán todos los cursos.")
        return resultados

    cursos_incluidos = resultados
    if keywords_incluir:
        # Generate regex patterns from inclusion keywords
        patrones_incluir = generar_patrones_regex(keywords_incluir)
        cursos_incluidos = [
            curso for curso in resultados
            if any(re.search(patron, curso['coursename']) for patron in patrones_incluir)
        ]

        # Identify unmatched keywords
        unmatched_keywords = [
            keyword for keyword in keywords_incluir
            if not any(re.search(rf"(?i).*{re.escape(keyword)}.*", curso['coursename']) for curso in resultados)
        ]
        if unmatched_keywords:
            print(f"Advertencia: No se encontraron cursos que coincidan con las palabras clave: {', '.join(unmatched_keywords)}")

    if keywords_excluir:
        # Generate regex patterns from exclusion keywords
        patrones_excluir = generar_patrones_regex(keywords_excluir)
        cursos_incluidos = [
            curso for curso in cursos_incluidos
            if not any(re.search(patron, curso['coursename']) for patron in patrones_excluir)
        ]

    if not cursos_incluidos:
        print("No se encontraron cursos que coincidan con las palabras clave en config.py.")
        print("¿Deseas descargar todos los cursos disponibles? (s/n):")
        respuesta = input("> ").strip().lower()
        if respuesta == 's':
            return resultados
        else:
            print("Saliendo del programa...")
            exit()

    return cursos_incluidos

# Function to allow the user to select courses interactively
def seleccionar_cursos(resultados):
    print("\nCursos disponibles:")
    for i, curso in enumerate(resultados, start=1):
        print(f"{i}. {curso['coursename']}")

    print("\nSelecciona los cursos que deseas descargar (separados por comas):")
    seleccion = input("Ejemplo: 1,3,5\n> ")
    indices = [int(x.strip()) - 1 for x in seleccion.split(",") if x.strip().isdigit()]
    seleccionados = [resultados[i] for i in indices if 0 <= i < len(resultados)]

    if not seleccionados:
        print("No se seleccionaron cursos válidos. Saliendo...")
        exit()

    return seleccionados

def main():
    # Verificar conexión a internet
    check_internet_connection()

    # 1. Iniciar el navegador con Chrome
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)

    try:
        # 2. Iniciar sesión
        login(driver, USUARIX, CONTRASEÑA)

        # 3. Obtener cursos
        resultados = obtener_cursos(driver)

        # 4. Filtrar cursos usando palabras clave y exclusiones
        resultados = filtrar_cursos_por_keywords(resultados, CURSOS_A_DESCARGAR, CURSOS_A_EXCLUIR)

        # 5. Descargar archivos y actualizar URLs
        registro = descargar_archivos_y_actualizar_urls(driver, resultados)

        # 6. Guardar el registro en un archivo Excel
        df_registro = pd.DataFrame(registro)
        excel_file = os.path.join(DIRECTORIO_DESCARGAS or os.getcwd(), 'registro_descargas.xlsx')
        guardar_registro_en_excel(df_registro, excel_file)

    finally:
        # 7. Borrar el archivo CSV
        csv_file = 'registro_descargas.csv'
        if os.path.exists(csv_file):
            os.remove(csv_file)
            print("Archivo CSV eliminado.")

        # 8. Borrar datos sensibles en config.py
        with open('config.py', 'r') as config_file:
            lines = config_file.readlines()

        with open('config.py', 'w') as config_file:
            for line in lines:
                if line.startswith("USUARIX") or line.startswith("CONTRASEÑA"):
                    if "USUARIX" in line:
                        config_file.write("USUARIX = ''  # Reemplazar con tu usuario\n")
                    elif "CONTRASEÑA" in line:
                        config_file.write("CONTRASEÑA = ''  # Reemplazar con tu contraseña\n")
                else:
                    config_file.write(line)
        print("Datos sensibles eliminados de config.py.")

        driver.quit()

if __name__ == "__main__":
    main()